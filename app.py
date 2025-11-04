"""
Ứng dụng chuyển đổi địa chỉ 2 cấp
Giao diện GUI bằng tiếng Việt cho người dùng - CustomTkinter version
"""
import customtkinter as ctk
from tkinter import filedialog, messagebox
import requests
import os
import threading
import time
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

TEST_API_KEY = "testing-chuyen-doi-2-cap"
API_DOMAIN_PROD = "https://address-converter.io.vn"
API_DOMAIN_TEST = "http://localhost:5000"
GOONG_API_KEY = "blank"
OPENMAP_API_KEY = "A5PFHXxiTe780Z2GZRQjPMjQ5EQaVCsm"

MAX_BATCH_SIZE = 1000
BATCH_DELAY_SECONDS = 5
OPENMAP_RATE_LIMIT_DELAY = 0.2

REPLACEMENT_REQUIRES = [
    ("Quận 02", "Thành phố Thủ Đức"),
    ("Quận 2", "Thành phố Thủ Đức"),
    ("Huyện Dăk RLấp", "Huyện Đắk R'Lấp"),
    ("Xã Ea Bhôk", "Xã Ea BHốk"),
    ("Huyện Dăk Mil", "Huyện Đắk Mil"),
    ("Thành phố KonTum" , "Thành phố Kon Tum"),
    ("Xã Đăk Ngo", "Xã Đắk Ngo")
]

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

def normalize_address_component(text):
    """
    Normalize address component by trimming whitespace and removing leading zeros.
    
    Examples:
        "Phường 02" -> "Phường 2"
        "Quận 03" -> "Quận 3"
        "  Phường 10  " -> "Phường 10"
    
    Args:
        text: Address component string
        
    Returns:
        Normalized string with trimmed spaces and no leading zeros
    """
    if not text:
        return ''
    
    text = text.strip()
    
    for old, new in REPLACEMENT_REQUIRES:
        if text.lower() == old.lower():
            return new
    
    text = text.replace("Đăk", "Đắk")
    text = text.replace("Dăk", "Đắk")
        
    parts = text.split()
    if len(parts) >= 2:
        last_part = parts[-1]
        if last_part.isdigit():
            normalized_number = str(int(last_part))
            parts[-1] = normalized_number
            return ' '.join(parts)
    
    return text


def remove_so_nha_prefix(text):
    """
    Remove 'số nhà' or 'Số nhà' prefix from street address and format properly.
    
    Examples:
        "số nhà 123 Đường ABC" -> "123, Đường ABC"
        "Số nhà 45" -> "45"
        "123 Đường XYZ" -> "123, Đường XYZ"
    
    Args:
        text: Street address string
        
    Returns:
        String with 'số nhà' prefix removed and properly formatted
    """
    if not text:
        return ''
    
    text = text.strip()
    
    if text.lower().startswith('số nhà'):
        remaining = text[len('số nhà'):].strip()
        if remaining.startswith(':'):
            remaining = remaining[1:].strip()
        parts = remaining.split(None, 1)
        if len(parts) == 2:
            return f"{parts[0]}, {parts[1]}"
        return remaining
    
    parts = text.split(None, 1)
    if len(parts) == 2 and parts[0].replace('.', '').replace('/', '').isdigit():
        return f"{parts[0]}, {parts[1]}"
    
    return text


def geocode_address(address, goong_api_key):
    """
    Call Goong.io Geocoding API to convert address to coordinates.
    
    Args:
        address: Full address string
        goong_api_key: Goong.io API key
        
    Returns:
        tuple: (success, latitude, longitude, error_message)
    """
    if not address or goong_api_key == "blank":
        return False, None, None, "Geocoding disabled"
    
    try:
        import urllib.parse
        encoded_address = urllib.parse.quote(address)
        geocode_url = f"https://rsapi.goong.io/geocode?address={encoded_address}&api_key={goong_api_key}"
        
        response = requests.get(geocode_url, timeout=10)
        
        if response.status_code != 200:
            return False, None, None, f"HTTP {response.status_code}"
        
        data = response.json()
        
        if data.get('status') != 'OK':
            return False, None, None, "No results found"
        
        results = data.get('results', [])
        if not results:
            return False, None, None, "Empty results"
        
        geometry = results[0].get('geometry', {})
        location = geometry.get('location', {})
        
        lat = location.get('lat')
        lng = location.get('lng')
        
        if lat is None or lng is None:
            return False, None, None, "No coordinates in response"
        
        return True, lat, lng, None
        
    except Exception:
        return False, None, None, "Exception"



def openmap_geocode_address(address, openmap_api_key):
    """
    Call OpenMap.vn Forward Geocoding API to convert address to coordinates.
    
    Uses Google-compatible format endpoint for consistency with existing code.
    
    Args:
        address: Full address string
        openmap_api_key: OpenMap.vn API key
        
    Returns:
        tuple: (success, latitude, longitude, error_message)
    """
    if not address or openmap_api_key == "blank":
        return False, None, None, "OpenMap geocoding disabled"
    
    try:
        import urllib.parse
        encoded_address = urllib.parse.quote(address)
        geocode_url = f"https://mapapis.openmap.vn/v1/geocode/forward?address={encoded_address}&apikey={openmap_api_key}"
        
        response = requests.get(geocode_url, timeout=10)
        
        if response.status_code != 200:
            return False, None, None, f"HTTP {response.status_code}"
        
        data = response.json()
        
        if data.get('status') != 'OK':
            return False, None, None, f"Status: {data.get('status')}"
        
        results = data.get('results', [])
        if not results:
            return False, None, None, "No results"
        
        geometry = results[0].get('geometry', {})
        location = geometry.get('location', {})
        
        lat = location.get('lat')
        lng = location.get('lng')
        
        if lat is None or lng is None:
            return False, None, None, "No coordinates"
        
        return True, lat, lng, None
        
    except requests.exceptions.Timeout:
        return False, None, None, "Timeout"
    except requests.exceptions.RequestException as e:
        return False, None, None, f"Request error: {str(e)}"
    except Exception as e:
        return False, None, None, f"Exception: {str(e)}"


def convert_coordinates_to_address(coordinates_list, api_key, api_domain):
    """
    Call /api/convert-coordinate to get new addresses from coordinates (batch).
    
    Args:
        coordinates_list: List of tuples [(longitude, latitude), ...]
        api_key: Address converter API key
        api_domain: API domain (test or prod)
        
    Returns:
        list: List of tuples [(success, new_address, error_message), ...]
    """
    if not coordinates_list:
        return []
    
    try:
        coordinate_endpoint = f"{api_domain}/api/convert-coordinate"
        
        coord_objects = [{"longitude": lng, "latitude": lat} for lng, lat in coordinates_list]
        
        response = requests.post(
            coordinate_endpoint,
            json={
                "coordinates": coord_objects,
                "key": api_key
            },
            headers={"Content-Type": "application/json"},
            timeout=60
        )
        
        if response.status_code != 200:
            error_msg = f"HTTP {response.status_code}"
            return [(False, None, error_msg) for _ in coordinates_list]
        
        result_data = response.json()
        
        if not result_data.get("success", False):
            error = result_data.get("error", "Unknown error")
            return [(False, None, error) for _ in coordinates_list]
        
        data_list = result_data.get("data", [])
        if not data_list:
            return [(False, None, "Empty data") for _ in coordinates_list]
        
        results = []
        for i, coord_result in enumerate(data_list):
            if not coord_result.get("success", False):
                error = coord_result.get("error", "Coordinate conversion failed")
                results.append((False, None, error))
                continue
            
            ward_info = coord_result.get("wardInfo", {})
            ward_name = ward_info.get("newWardName", "")
            province_name = ward_info.get("provinceName", "")
            
            if ward_name and province_name:
                new_address = f"{ward_name}, {province_name}"
                results.append((True, new_address, None))
            else:
                results.append((False, None, "Incomplete ward info"))
        
        while len(results) < len(coordinates_list):
            results.append((False, None, "Missing result"))
        
        return results
        
    except Exception as e:
        error_msg = f"Exception: {str(e)}"
        return [(False, None, error_msg) for _ in coordinates_list]


def try_geocoding_fallback(original_address, row_data, is_multi_column, api_key, api_domain, goong_api_key, openmap_api_key):
    """
    Try to convert failed address using geocoding fallback.
    
    Process:
    1. For multi-column: prepare address from street, ward, district, province
    2. Try Goong.io geocoding API first, if unavailable try OpenMap.vn API
    3. Call coordinate conversion API to get new address
    
    Args:
        original_address: Original address string that failed
        row_data: Original row data dict (for multi-column)
        is_multi_column: Boolean indicating multi-column format
        api_key: Address converter API key
        api_domain: API domain
        goong_api_key: Goong.io API key
        openmap_api_key: OpenMap.vn API key
        
    Returns:
        dict: Result with success, converted, error fields
    """
    if goong_api_key == "blank" and openmap_api_key == "blank":
        return {
            "original": original_address,
            "converted": "",
            "error": "Geocoding disabled",
            "success": False
        }
    
    geocode_address_str = original_address
    
    if is_multi_column and row_data:
        so_nha_duong = row_data.get('so_nha_duong', '').strip()
        phuong_xa = row_data.get('phuong_xa', '').strip()
        quan_huyen = row_data.get('quan_huyen', '').strip()
        tinh_thanh = row_data.get('tinh_thanh', '').strip()
        
        so_nha_duong = remove_so_nha_prefix(so_nha_duong)
        
        addr_parts = [p for p in [so_nha_duong, phuong_xa, quan_huyen, tinh_thanh] if p]
        geocode_address_str = ', '.join(addr_parts)
    
    geo_success = False
    lat = None
    lng = None
    geo_error = None
    
    if goong_api_key != "blank":
        geo_success, lat, lng, geo_error = geocode_address(geocode_address_str, goong_api_key)
    
    if not geo_success and openmap_api_key != "blank":
        geo_success, lat, lng, geo_error = openmap_geocode_address(geocode_address_str, openmap_api_key)
    
    if not geo_success:
        return {
            "original": original_address,
            "converted": "",
            "error": "Geocoding failed",
            "success": False
        }
    
    coord_results = convert_coordinates_to_address([(lng, lat)], api_key, api_domain)
    
    if coord_results and len(coord_results) > 0:
        coord_success, new_address, coord_error = coord_results[0]
        
        if coord_success:
            return {
                "original": original_address,
                "converted": new_address,
                "success": True,
                "geocoded": True
            }
        else:
            combined_error = f"Geocode OK but coordinate conversion failed: {coord_error}"
            return {
                "original": original_address,
                "converted": "",
                "error": combined_error,
                "success": False
            }
    else:
        return {
            "original": original_address,
            "converted": "",
            "error": "Coordinate conversion returned no results",
            "success": False
        }


class AddressConverterApp:
    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("Chuyển đổi địa chỉ 2 cấp")
        self.root.geometry("800x750")
        
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        
        self.api_key_var = ctk.StringVar()
        self.input_file_path = ctk.StringVar()
        self.output_folder = ""
        self.is_converting = False
        self.stop_requested = False
        self.api_key_visible = False
        
        self.last_openmap_request_time = 0
        
        self.account_info = {
            'name': '',
            'balance': 0,
            'createdAt': '',
            'updatedAt': ''
        }
        
        self.create_widgets()
        
    def create_widgets(self):
        main_frame = ctk.CTkFrame(self.root, corner_radius=15)
        main_frame.grid(row=0, column=0, sticky="nsew", padx=15, pady=15)
        main_frame.grid_columnconfigure(1, weight=1)
        
        title_label = ctk.CTkLabel(
            main_frame,
            text="CHUYỂN ĐỔI ĐỊA CHỈ 2 CẤP",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.grid(row=0, column=0, columnspan=3, pady=(20, 30), padx=20)
        
        ctk.CTkLabel(
            main_frame,
            text="API Key:",
            font=ctk.CTkFont(size=13)
        ).grid(row=1, column=0, sticky="w", padx=20, pady=(10, 5))
        
        api_key_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        api_key_frame.grid(row=1, column=1, columnspan=2, sticky="ew", padx=20, pady=(10, 5))
        api_key_frame.grid_columnconfigure(0, weight=1)
        
        self.api_entry = ctk.CTkEntry(
            api_key_frame,
            textvariable=self.api_key_var,
            width=500,
            height=35,
            placeholder_text="Nhập API Key của bạn...",
            show="•"
        )
        self.api_entry.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        self.api_entry.bind("<FocusOut>", lambda e: self.on_api_key_focus_out())
        self.api_entry.bind("<KeyRelease>", lambda e: self.auto_trim_api_key())
        
        self.toggle_api_btn = ctk.CTkButton(
            api_key_frame,
            text="Hiện",
            command=self.toggle_api_key_visibility,
            width=60,
            height=35,
            font=ctk.CTkFont(size=12)
        )
        self.toggle_api_btn.grid(row=0, column=1)
        
        self.account_info_frame = ctk.CTkFrame(
            main_frame,
            corner_radius=10,
            fg_color=("#e8f4f8", "#1a3a4a")
        )
        self.account_info_frame.grid(row=2, column=0, columnspan=3, sticky="ew", padx=20, pady=(10, 5))
        self.account_info_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        self.account_name_label = ctk.CTkLabel(
            self.account_info_frame,
            text="Tài khoản: --",
            font=ctk.CTkFont(size=11),
            anchor="w"
        )
        self.account_name_label.grid(row=0, column=0, sticky="w", padx=10, pady=8)
        
        self.account_balance_label = ctk.CTkLabel(
            self.account_info_frame,
            text="Số dư: --",
            font=ctk.CTkFont(size=11),
            anchor="w"
        )
        self.account_balance_label.grid(row=0, column=1, sticky="w", padx=10, pady=8)
        
        self.account_created_label = ctk.CTkLabel(
            self.account_info_frame,
            text="Ngày tạo: --",
            font=ctk.CTkFont(size=11),
            anchor="w"
        )
        self.account_created_label.grid(row=0, column=2, sticky="w", padx=10, pady=8)
        
        self.account_updated_label = ctk.CTkLabel(
            self.account_info_frame,
            text="Cập nhật: --",
            font=ctk.CTkFont(size=11),
            anchor="w"
        )
        self.account_updated_label.grid(row=0, column=3, sticky="w", padx=10, pady=8)
        
        self.account_info_frame.grid_remove()
        
        ctk.CTkLabel(
            main_frame,
            text="File đầu vào:",
            font=ctk.CTkFont(size=13)
        ).grid(row=3, column=0, sticky="w", padx=20, pady=(10, 5))
        
        self.input_entry = ctk.CTkEntry(
            main_frame,
            textvariable=self.input_file_path,
            width=380,
            height=35,
            placeholder_text="Chọn file .txt hoặc .xlsx...",
            state="readonly"
        )
        self.input_entry.grid(row=3, column=1, sticky="ew", padx=(20, 10), pady=(10, 5))
        
        self.browse_btn = ctk.CTkButton(
            main_frame,
            text="Chọn file",
            command=self.browse_input_file,
            width=100,
            height=35,
            font=ctk.CTkFont(size=13)
        )
        self.browse_btn.grid(row=3, column=2, sticky="e", padx=20, pady=(10, 5))
        
        info_label = ctk.CTkLabel(
            main_frame,
            text="(Hỗ trợ file .xlsx và .txt)",
            font=ctk.CTkFont(size=11),
            text_color="gray"
        )
        info_label.grid(row=4, column=1, sticky="w", padx=20, pady=(0, 10))
        
        button_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        button_frame.grid(row=5, column=0, columnspan=3, pady=20)
        
        self.convert_btn = ctk.CTkButton(
            button_frame,
            text="CHUYỂN ĐỔI",
            command=self.start_conversion,
            width=180,
            height=45,
            font=ctk.CTkFont(size=16, weight="bold"),
            fg_color=("#1f6aa5", "#144870"),
            hover_color=("#144870", "#0d3147")
        )
        self.convert_btn.grid(row=0, column=1, padx=10)
        
        self.stop_btn = ctk.CTkButton(
            button_frame,
            text="DỪNG",
            command=self.stop_conversion,
            width=180,
            height=45,
            font=ctk.CTkFont(size=16, weight="bold"),
            fg_color=("#d32f2f", "#b71c1c"),
            hover_color=("#b71c1c", "#8b0000"),
            state="disabled"
        )
        self.stop_btn.grid(row=0, column=0, padx=10)
        
        self.open_folder_btn = ctk.CTkButton(
            button_frame,
            text="MỞ THƯ MỤC KẾT QUẢ",
            command=self.open_output_folder,
            width=180,
            height=45,
            font=ctk.CTkFont(size=13),
            state="disabled"
        )
        self.open_folder_btn.grid(row=0, column=2, padx=10)
        
        separator = ctk.CTkFrame(main_frame, height=2, fg_color="gray30")
        separator.grid(row=6, column=0, columnspan=3, sticky="ew", padx=20, pady=(10, 20))
        
        self.progress_label = ctk.CTkLabel(
            main_frame,
            text="",
            font=ctk.CTkFont(size=12),
            text_color=("#1f6aa5", "#5fb4ff")
        )
        self.progress_label.grid(row=7, column=0, columnspan=3, pady=(10, 5))
        
        self.progress_bar = ctk.CTkProgressBar(
            main_frame,
            width=700,
            height=15,
            corner_radius=8,
            mode="determinate"
        )
        self.progress_bar.grid(row=8, column=0, columnspan=3, padx=20, pady=(0, 20))
        self.progress_bar.set(0)
        
        ctk.CTkLabel(
            main_frame,
            text="Kết quả:",
            font=ctk.CTkFont(size=14, weight="bold")
        ).grid(row=9, column=0, columnspan=3, sticky="w", padx=20, pady=(10, 10))
        
        self.result_textbox = ctk.CTkTextbox(
            main_frame,
            width=700,
            height=165,
            font=ctk.CTkFont(family="Consolas", size=11),
            wrap="word"
        )
        self.result_textbox.grid(row=10, column=0, columnspan=3, sticky="ew", padx=20, pady=(0, 10))
        self.result_textbox.configure(state="disabled")
        
    def browse_input_file(self):
        file_path = filedialog.askopenfilename(
            title="Chọn file đầu vào",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("Text files", "*.txt"),
                ("All files", "*.*")
            ]
        )
        if file_path:
            self.input_file_path.set(file_path)
            
    def update_result_text(self, message, clear=False):
        self.result_textbox.configure(state="normal")
        if clear:
            self.result_textbox.delete("0.0", "end")
        self.result_textbox.insert("end", message + "\n")
        self.result_textbox.see("end")
        self.result_textbox.configure(state="disabled")
    
    def update_progress(self, current, total, status=""):
        if total > 0:
            percentage = (current / total) * 100
            self.progress_bar.set(percentage / 100)
            
            if status:
                label_text = f"{status} - {current}/{total} ({percentage:.1f}%)"
            else:
                label_text = f"{current}/{total} ({percentage:.1f}%)"
            
            self.progress_label.configure(text=label_text)
        else:
            self.progress_bar.set(0)
            self.progress_label.configure(text=status)

    def auto_trim_api_key(self):
        current_value = self.api_key_var.get()
        trimmed_value = current_value.strip()
        if current_value != trimmed_value:
            cursor_pos = self.api_entry.index("insert")
            self.api_key_var.set(trimmed_value)
            self.api_entry.icursor(max(0, cursor_pos - (len(current_value) - len(trimmed_value))))
        
    def start_conversion(self):
        if self.is_converting:
            messagebox.showwarning("Cảnh báo", "Đang xử lý, vui lòng đợi!")
            return
            
        api_key = self.api_key_var.get().strip()
        input_file = self.input_file_path.get().strip()
        
        if not api_key:
            messagebox.showerror("Lỗi", "Vui lòng nhập API Key!")
            return
            
        if not input_file:
            messagebox.showerror("Lỗi", "Vui lòng chọn file đầu vào!")
            return
            
        if not os.path.exists(input_file):
            messagebox.showerror("Lỗi", "File không tồn tại!")
            return
            
        file_ext = Path(input_file).suffix.lower()
        if file_ext not in ['.txt', '.xlsx']:
            messagebox.showerror("Lỗi", "Chỉ hỗ trợ file .txt hoặc .xlsx!")
            return
        
        self.is_converting = True
        self.stop_requested = False
        self.convert_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.open_folder_btn.configure(state="disabled")
        self.root.after(0, lambda: self.update_progress(0, 100, "Đang khởi tạo..."))
        
        thread = threading.Thread(target=self.perform_conversion, args=(api_key, input_file, file_ext))
        thread.daemon = True
        thread.start()

    def stop_conversion(self):
        """Request to stop the conversion process"""
        if self.is_converting:
            self.stop_requested = True
            self.update_result_text("\n[DỪNG] Đang dừng quá trình chuyển đổi...")
            self.update_result_text("Sẽ lưu kết quả đã xử lý...")
            self.stop_btn.configure(state="disabled")
        
    def perform_conversion(self, api_key, input_file, file_ext):
        try:
            self.update_result_text("Đang đọc file đầu vào...", clear=True)
            self.root.after(0, lambda: self.update_progress(10, 100, "Đọc file đầu vào..."))
            
            if file_ext == '.txt':
                data, is_multi_column, headers = self.read_txt_file(input_file)
            else:
                data, is_multi_column, headers = self.read_excel_file(input_file)
            
            if not data:
                self.root.after(0, lambda: messagebox.showerror("Lỗi", "File đầu vào không có dữ liệu!"))
                return
            
            if is_multi_column:
                addresses = []
                original_data = data
                for row in data:
                    addr_parts = [
                        normalize_address_component(row.get('phuong_xa', '')),
                        normalize_address_component(row.get('quan_huyen', '')),
                        normalize_address_component(row.get('tinh_thanh', ''))
                    ]
                    address = ', '.join([p for p in addr_parts if p])
                    addresses.append(address)
                
                self.update_result_text(f"[ĐỊA CHỈ NHIỀU CỘT] Đã phát hiện định dạng {len(headers)} cột")
                self.update_result_text(f"Sử dụng pattern: Phường Xã, Quận Huyện, Tỉnh Thành")
            else:
                addresses = data
                original_data = None
                self.update_result_text(f"[ĐỊA CHỈ ĐƠN] Đã phát hiện định dạng 1 cột")
                
            total_addresses = len(addresses)
            self.update_result_text(f"Đã đọc {total_addresses} địa chỉ từ file")
            self.root.after(0, lambda: self.update_progress(20, 100, "Chuẩn bị xử lý..."))
            
            num_batches = (total_addresses + MAX_BATCH_SIZE - 1) // MAX_BATCH_SIZE
            
            if num_batches > 1:
                self.update_result_text(f"\n[CẢNH BÁO] File có {total_addresses} địa chỉ, vượt quá giới hạn {MAX_BATCH_SIZE} địa chỉ/lần")
                self.update_result_text(f"Hệ thống sẽ tự động chia thành {num_batches} lần gửi")
                self.update_result_text(f"Thời gian chờ giữa các lần: {BATCH_DELAY_SECONDS} giây\n")
            
            api_domain = API_DOMAIN_TEST if api_key == TEST_API_KEY else API_DOMAIN_PROD
            convert_endpoint = f"{api_domain}/api/convert-batch"

            results_dict = {}
            total_successful = 0
            total_failed = 0
            skipped_indices = set()
            
            for batch_num in range(num_batches):
                if self.stop_requested:
                    self.update_result_text("\n[DỪNG] Đã dừng tại batch {}/{}".format(batch_num + 1, num_batches))
                    break
                
                start_idx = batch_num * MAX_BATCH_SIZE
                end_idx = min((batch_num + 1) * MAX_BATCH_SIZE, total_addresses)
                
                batch_items = [(i, addresses[i]) for i in range(start_idx, end_idx) if i not in skipped_indices]
                
                batch_info = f"Lần {batch_num + 1}/{num_batches}" if num_batches > 1 else ""
                if batch_info:
                    self.update_result_text(f"{batch_info}: Xử lý địa chỉ {start_idx + 1} đến {end_idx}")
                
                retry_count = 0
                max_retries = 10000
                
                while retry_count < max_retries and batch_items and not self.stop_requested:
                    progress_start = 20 + (batch_num * 70 // num_batches)
                    retry_info = f" (Retry {retry_count})" if retry_count > 0 else ""
                    self.root.after(0, lambda p=progress_start, b=batch_info, r=retry_info: self.update_progress(
                        p, 100, f"Đang gửi yêu cầu {b}{r}..." if b else f"Đang gửi yêu cầu{r}..."
                    ))
                    
                    batch_addresses = [addr for idx, addr in batch_items]
                    
                    response = requests.post(
                        convert_endpoint,
                        json={
                            "addresses": batch_addresses,
                            "key": api_key
                        },
                        headers={"Content-Type": "application/json"},
                        timeout=60
                    )
                    
                    result_data = response.json()
                    
                    if not result_data.get("success", False):
                        error_msg = result_data.get("error", "Lỗi không xác định")
                        
                        if result_data.get("rateLimited", False):
                            retry_after = result_data.get("retryAfter", 0)
                            self.update_result_text(f"\n[RATE LIMIT] {error_msg}")
                            self.update_result_text(f"Tự động thử lại sau {retry_after} giây...")
                            
                            for i in range(retry_after):
                                if self.stop_requested:
                                    break
                                time.sleep(1)
                                remaining = retry_after - i - 1
                                if remaining > 0:
                                    self.root.after(0, lambda r=remaining: self.progress_label.configure(
                                        text=f"Rate limited - Chờ {r} giây..."
                                    ))
                            if self.stop_requested:
                                break
                            continue
                        else:
                            self.update_result_text(f"\n[LỖI] {error_msg}")
                            self.root.after(0, lambda e=error_msg: messagebox.showerror("Lỗi API", e))
                            return
                    
                    data = result_data.get("data", {})
                    batch_results = data.get("results", [])
                    
                    balance_error_indices = []
                    balance_error_addrs = set()
                    
                    for i, result in enumerate(batch_results):
                        original_idx = batch_items[i][0]
                        
                        if not result.get("success", False):
                            error_msg = result.get("error", "")
                            if "Số dư không đủ" in error_msg:
                                balance_error_indices.append(original_idx)
                                balance_error_addrs.add(batch_items[i][1])
                                results_dict[original_idx] = {
                                    "original": result.get("original", ""),
                                    "converted": "",
                                    "error": "Số dư không đủ - Đã bỏ qua",
                                    "success": False
                                }
                            else:
                                results_dict[original_idx] = result
                                total_failed += 1
                        else:
                            results_dict[original_idx] = result
                            total_successful += 1
                    
                    if balance_error_indices:
                        skipped_indices.update(balance_error_indices)
                        
                        duplicates_in_remaining = 0
                        for idx, addr in batch_items:
                            if idx not in balance_error_indices and addr in balance_error_addrs:
                                skipped_indices.add(idx)
                                results_dict[idx] = {
                                    "original": addr,
                                    "converted": "",
                                    "error": "Số dư không đủ - Đã bỏ qua (trùng)",
                                    "success": False
                                }
                                duplicates_in_remaining += 1
                        
                        total_skipped = len(balance_error_indices)
                        self.update_result_text(f"\n[SỐ DƯ KHÔNG ĐỦ] Bỏ qua {total_skipped} địa chỉ")
                        if duplicates_in_remaining > 0:
                            self.update_result_text(f"  + Bỏ qua thêm {duplicates_in_remaining} địa chỉ trùng")
                        
                        batch_items = [(idx, addr) for idx, addr in batch_items if idx not in skipped_indices]
                        
                        if batch_items:
                            self.update_result_text(f"Còn lại {len(batch_items)} địa chỉ, tiếp tục xử lý...")
                            retry_count += 1
                            if not self.stop_requested:
                                time.sleep(1)
                            continue
                        else:
                            self.update_result_text("Tất cả địa chỉ trong batch đã được xử lý")
                            break
                    
                    break
                
                if retry_count >= max_retries:
                    self.update_result_text(f"\n[CẢNH BÁO] Đã vượt quá số lần retry tối đa ({max_retries})")
                
                if batch_num < num_batches - 1:
                    if self.stop_requested:
                        break
                    self.update_result_text(f"Chờ {BATCH_DELAY_SECONDS} giây trước khi gửi lần tiếp theo...")
                    for i in range(BATCH_DELAY_SECONDS):
                        if self.stop_requested:
                            break
                        time.sleep(1)
                        remaining = BATCH_DELAY_SECONDS - i - 1
                        if remaining > 0:
                            self.root.after(0, lambda r=remaining: self.progress_label.configure(
                                text=f"Chờ {r} giây..."
                            ))
            
            for i in range(total_addresses):
                if i not in results_dict:
                    results_dict[i] = {
                        "original": addresses[i],
                        "converted": "",
                        "error": "Chưa xử lý (đã dừng)" if self.stop_requested else "Chưa xử lý",
                        "success": False
                    }
            
            if not self.stop_requested and (GOONG_API_KEY != "blank" or OPENMAP_API_KEY != "blank"):
                self.update_result_text(f"\n{'='*50}")
                self.update_result_text("[TRY] Đang thử geocoding cho địa chỉ lỗi...")
                self.root.after(0, lambda: self.update_progress(85, 100, "Geocoding fallback..."))
                
                failed_indices = [i for i in range(total_addresses) 
                                if not results_dict[i].get("success", False) 
                                and i not in skipped_indices]
                
                if failed_indices:
                    self.update_result_text(f"Tìm thấy {len(failed_indices)} địa chỉ lỗi, đang thử geocoding...")
                    
                    geocode_tasks = []
                    for idx in failed_indices:
                        if self.stop_requested:
                            break
                        
                        geocode_address_str = addresses[idx]
                        row_data = None
                        
                        if is_multi_column and original_data:
                            row_data = original_data[idx]
                            so_nha_duong = row_data.get('so_nha_duong', '').strip()
                            phuong_xa = row_data.get('phuong_xa', '').strip()
                            quan_huyen = row_data.get('quan_huyen', '').strip()
                            tinh_thanh = row_data.get('tinh_thanh', '').strip()
                            
                            so_nha_duong = remove_so_nha_prefix(so_nha_duong)
                            
                            addr_parts = [p for p in [so_nha_duong, phuong_xa, quan_huyen, tinh_thanh] if p]
                            geocode_address_str = ', '.join(addr_parts)
                        
                        geocode_tasks.append((idx, geocode_address_str, addresses[idx]))
                    
                    coordinates_to_convert = []
                    geocode_map = {}
                    
                    for idx, geocode_addr, original_addr in geocode_tasks:
                        geo_success = False
                        lat = None
                        lng = None
                        
                        if GOONG_API_KEY != "blank":
                            geo_success, lat, lng, geo_error = geocode_address(geocode_addr, GOONG_API_KEY)
                        
                        if not geo_success and OPENMAP_API_KEY != "blank":
                            geo_success, lat, lng, geo_error = self.openmap_geocode_with_rate_limit(geocode_addr, OPENMAP_API_KEY)
                        
                        if geo_success:
                            coord_idx = len(coordinates_to_convert)
                            coordinates_to_convert.append((lng, lat))
                            geocode_map[coord_idx] = (idx, original_addr)
                    
                    self.update_result_text(f"Geocoding thành công: {len(coordinates_to_convert)}/{len(geocode_tasks)} địa chỉ")
                    
                    geocoded_success = 0
                    geocoded_failed = len(geocode_tasks) - len(coordinates_to_convert)
                    
                    if coordinates_to_convert:
                        self.update_result_text(f"Đang chuyển đổi {len(coordinates_to_convert)} tọa độ...")
                        
                        coord_results = convert_coordinates_to_address(
                            coordinates_to_convert,
                            api_key,
                            api_domain
                        )
                        
                        for coord_idx, (success, new_address, error) in enumerate(coord_results):
                            if coord_idx in geocode_map:
                                idx, original_addr = geocode_map[coord_idx]
                                
                                if success:
                                    results_dict[idx] = {
                                        "original": original_addr,
                                        "converted": new_address,
                                        "success": True,
                                        "geocoded": True
                                    }
                                    geocoded_success += 1
                                    total_successful += 1
                                    total_failed -= 1
                                else:
                                    results_dict[idx] = {
                                        "original": original_addr,
                                        "converted": "",
                                        "error": f"Geocode OK but coordinate conversion failed: {error}",
                                        "success": False
                                    }
                                    geocoded_failed += 1
                    
                    self.update_result_text(f"Kết quả geocoding: {geocoded_success} thành công, {geocoded_failed} thất bại")
                else:
                    self.update_result_text("Không có địa chỉ lỗi nào cần geocoding")
            
            all_results = [results_dict[i] for i in range(total_addresses)]
            
            self.root.after(0, lambda: self.update_progress(90, 100, "Đang lưu kết quả..."))

            total_skipped = len(skipped_indices)
            total_not_processed = sum(1 for r in all_results if not r.get("success", False) and "Chưa xử lý" in r.get("error", ""))
            
            self.update_result_text(f"\n{'='*50}")
            if self.stop_requested:
                self.update_result_text(f"ĐÃ DỪNG - TỔNG KẾT:")
            else:
                self.update_result_text(f"TỔNG KẾT:")
            self.update_result_text(f"  - Tổng số địa chỉ: {total_addresses}")
            self.update_result_text(f"  - Thành công: {total_successful}")
            self.update_result_text(f"  - Thất bại: {total_failed}")
            if total_skipped > 0:
                self.update_result_text(f"  - Bỏ qua (Số dư không đủ): {total_skipped}")
            if total_not_processed > 0:
                self.update_result_text(f"  - Chưa xử lý: {total_not_processed}")
            if num_batches > 1:
                self.update_result_text(f"  - Số lần gửi: {num_batches}")
            self.update_result_text(f"{'='*50}\n")
            
            input_path = Path(input_file)
            output_folder = input_path.parent / "output"
            output_folder.mkdir(exist_ok=True)
            self.output_folder = str(output_folder)
            
            output_filename = f"{input_path.stem}_converted{file_ext}"
            output_path = output_folder / output_filename
            
            if file_ext == '.txt':
                self.write_txt_output(output_path, all_results, is_multi_column, headers, original_data)
            else:
                self.write_excel_output(output_path, all_results, is_multi_column, headers, original_data)
            
            self.update_result_text(f"Đã lưu kết quả tại:\n{output_path}")
            if self.stop_requested:
                self.root.after(0, lambda: self.update_progress(100, 100, "Đã dừng - Đã lưu kết quả!"))
            else:
                self.root.after(0, lambda: self.update_progress(100, 100, "Hoàn thành!"))
            
            self.root.after(0, lambda: self.open_folder_btn.configure(state="normal"))
            
            if self.stop_requested:
                self.root.after(0, lambda s=total_successful, f=total_failed, sk=total_skipped, np=total_not_processed, fn=output_filename: messagebox.showinfo(
                    "Đã dừng", 
                    f"Đã dừng và lưu kết quả!\n\nThành công: {s}\nThất bại: {f}\nBỏ qua: {sk}\nChưa xử lý: {np}\n\nFile kết quả: {fn}"
                ))
            else:
                self.root.after(0, lambda s=total_successful, f=total_failed, sk=total_skipped, fn=output_filename: messagebox.showinfo(
                    "Hoàn thành", 
                    f"Chuyển đổi thành công!\n\nThành công: {s}\nThất bại: {f}\nBỏ qua: {sk}\n\nFile kết quả: {fn}"
                ))
            
        except requests.exceptions.Timeout:
            self.update_result_text("\n[LỖI] Timeout - Không thể kết nối đến API")
            self.root.after(0, lambda: messagebox.showerror("Lỗi", "Không thể kết nối đến API (Timeout)"))
        except requests.exceptions.ConnectionError:
            self.update_result_text("\n[LỖI] Không thể kết nối đến server")
            if api_key == TEST_API_KEY:
                self.update_result_text("Hãy chắc chắn test server đang chạy tại http://localhost:5000")
            self.root.after(0, lambda: messagebox.showerror("Lỗi", "Không thể kết nối đến server"))
        except Exception as e:
            self.update_result_text(f"\n[LỖI] {str(e)}")
            self.root.after(0, lambda e=str(e): messagebox.showerror("Lỗi", f"Lỗi xử lý: {e}"))
        finally:
            self.is_converting = False
            self.stop_requested = False
            self.root.after(0, lambda: self.convert_btn.configure(state="normal"))
            self.root.after(0, lambda: self.stop_btn.configure(state="disabled"))
            self.root.after(0, lambda: self.progress_label.configure(text=""))
            
    def read_txt_file(self, file_path):
        """Read txt file and detect format: single column or multi-column (5+ columns)
        
        Returns:
            tuple: (addresses_or_rows, is_multi_column, headers)
            - addresses_or_rows: list of strings (single column) or list of dicts (multi-column)
            - is_multi_column: boolean indicating format type
            - headers: list of header names (for multi-column) or None
        """
        data = []
        is_multi_column = False
        headers = None
        
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = [line.strip() for line in f if line.strip()]
            
            if not lines:
                return [], False, None
            
            first_line_parts = [p.strip() for p in lines[0].split(',')]
            
            if len(first_line_parts) >= 5:
                is_multi_column = True
                headers = first_line_parts
                
                for line in lines[1:]:
                    parts = [p.strip() for p in line.split(',')]
                    if len(parts) >= 5:
                        row_data = {
                            'code': parts[0],
                            'so_nha_duong': parts[1],
                            'phuong_xa': parts[2],
                            'quan_huyen': parts[3],
                            'tinh_thanh': parts[4],
                            'extra_columns': parts[5:] if len(parts) > 5 else []
                        }
                        data.append(row_data)
            else:
                for line in lines:
                    if line:
                        data.append(line)
        
        except Exception as e:
            self.update_result_text(f"[LỖI] Không thể đọc file txt: {str(e)}")
            return [], False, None
        
        return data, is_multi_column, headers
        
    def read_excel_file(self, file_path):
        """Read Excel file and detect format: single column or multi-column (5+ columns)
        
        Returns:
            tuple: (addresses_or_rows, is_multi_column, headers)
            - addresses_or_rows: list of strings (single column) or list of dicts (multi-column)
            - is_multi_column: boolean indicating format type
            - headers: list of header names (for multi-column) or None
        """
        data = []
        is_multi_column = False
        headers = None
        
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active
            
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            
            if not rows:
                return [], False, None
            
            first_row = rows[0]
            num_cols = sum(1 for cell in first_row if cell is not None and str(cell).strip())
            
            if num_cols >= 5:
                is_multi_column = True
                headers = [str(cell).strip() if cell is not None else '' for cell in first_row]
                
                for row in rows[1:]:
                    if row and any(cell is not None for cell in row):
                        cells = [str(cell).strip() if cell is not None else '' for cell in row]
                        if len(cells) >= 5 and any(cells[:5]):
                            row_data = {
                                'code': cells[0] if len(cells) > 0 else '',
                                'so_nha_duong': cells[1] if len(cells) > 1 else '',
                                'phuong_xa': cells[2] if len(cells) > 2 else '',
                                'quan_huyen': cells[3] if len(cells) > 3 else '',
                                'tinh_thanh': cells[4] if len(cells) > 4 else '',
                                'extra_columns': cells[5:] if len(cells) > 5 else []
                            }
                            data.append(row_data)
            else:
                for row in rows:
                    if row and row[0]:
                        addr = str(row[0]).strip()
                        if addr:
                            data.append(addr)
        
        except Exception as e:
            self.update_result_text(f"[LỖI] Không thể đọc file Excel: {str(e)}")
            return [], False, None
        
        return data, is_multi_column, headers
        
    def write_txt_output(self, output_path, results, is_multi_column=False, headers=None, original_data=None):
        """Write output to txt file
        
        Args:
            output_path: Path to output file
            results: List of conversion results
            is_multi_column: Boolean indicating if multi-column format
            headers: List of header names (for multi-column)
            original_data: List of original row dicts (for multi-column)
        """
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                if is_multi_column and headers and original_data:
                    new_headers = headers + ['Phường/ Xã mới', 'Tỉnh/ Thành mới']
                    f.write(','.join(new_headers) + '\n')
                    
                    for idx, result in enumerate(results):
                        if idx < len(original_data):
                            row_data = original_data[idx]
                            row_parts = [
                                row_data.get('code', ''),
                                row_data.get('so_nha_duong', ''),
                                row_data.get('phuong_xa', ''),
                                row_data.get('quan_huyen', ''),
                                row_data.get('tinh_thanh', '')
                            ]
                            row_parts.extend(row_data.get('extra_columns', []))
                            
                            if result.get("success", False):
                                converted = result.get("converted", "")
                                parts = [p.strip() for p in converted.split(',')]
                                ward = parts[0] if len(parts) > 0 else ""
                                province = parts[1] if len(parts) > 1 else ""
                            else:
                                error_msg = result.get("error", "Lỗi không xác định")
                                ward = f"LỖI: {error_msg}"
                                province = ""
                            

                            row_parts.append(ward)
                            row_parts.append(province)
                            f.write(','.join(row_parts) + '\n')
                else:
                    for result in results:
                        if result.get("success", False):
                            f.write(result.get("converted", "") + "\n")
                        else:
                            error_msg = result.get("error", "Lỗi không xác định")
                            f.write(f"LỖI: {error_msg}\n")
        except Exception as e:
            self.update_result_text(f"[LỖI] Không thể ghi file txt: {str(e)}")
            
    def write_excel_output(self, output_path, results, is_multi_column=False, headers=None, original_data=None):
        """Write output to Excel file
        
        Args:
            output_path: Path to output file
            results: List of conversion results
            is_multi_column: Boolean indicating if multi-column format
            headers: List of header names (for multi-column)
            original_data: List of original row dicts (for multi-column)
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Kết quả"
            
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            if is_multi_column and headers and original_data:
                new_headers = headers + ['Phường/ Xã mới', 'Tỉnh/ Thành mới']
                
                for col_idx, header in enumerate(new_headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                for idx, result in enumerate(results, start=2):
                    orig_idx = idx - 2
                    if orig_idx < len(original_data):
                        row_data = original_data[orig_idx]
                        
                        ws.cell(row=idx, column=1, value=row_data.get('code', ''))
                        ws.cell(row=idx, column=2, value=row_data.get('so_nha_duong', ''))
                        ws.cell(row=idx, column=3, value=row_data.get('phuong_xa', ''))
                        ws.cell(row=idx, column=4, value=row_data.get('quan_huyen', ''))
                        ws.cell(row=idx, column=5, value=row_data.get('tinh_thanh', ''))
                        
                        extra_cols = row_data.get('extra_columns', [])
                        for extra_idx, extra_val in enumerate(extra_cols):
                            ws.cell(row=idx, column=6+extra_idx, value=extra_val)
                        
                        ward_col = 6 + len(extra_cols)
                        province_col = ward_col + 1
                        
                        if result.get("success", False):
                            converted = result.get("converted", "")
                            parts = [p.strip() for p in converted.split(',')]
                            ward = parts[0] if len(parts) > 0 else ""
                            province = parts[1] if len(parts) > 1 else ""
                            
                            ws.cell(row=idx, column=ward_col, value=ward)
                            ws.cell(row=idx, column=province_col, value=province)
                        else:
                            error_msg = result.get("error", "Lỗi không xác định")
                            cell = ws.cell(row=idx, column=ward_col, value=f"LỖI: {error_msg}")
                            cell.font = Font(color="FF0000")
                            ws.cell(row=idx, column=province_col, value="")
                
                for col_idx in range(1, len(new_headers) + 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20
            else:
                ws['A1'] = "Gốc"
                ws['B1'] = "Chuyển đổi"
                
                ws['A1'].font = header_font
                ws['B1'].font = header_font
                ws['A1'].fill = header_fill
                ws['B1'].fill = header_fill
                ws['A1'].alignment = header_alignment
                ws['B1'].alignment = header_alignment
                
                ws.column_dimensions['A'].width = 50
                ws.column_dimensions['B'].width = 60
                
                for idx, result in enumerate(results, start=2):
                    ws[f'A{idx}'] = result.get("original", "")
                    if result.get("success", False):
                        ws[f'B{idx}'] = result.get("converted", "")
                    else:
                        error_msg = result.get("error", "Lỗi không xác định")
                        ws[f'B{idx}'] = f"LỖI: {error_msg}"
                        ws[f'B{idx}'].font = Font(color="FF0000")
            
            wb.save(output_path)
            wb.close()
        except Exception as e:
            self.update_result_text(f"[LỖI] Không thể ghi file Excel: {str(e)}")
            
    def open_output_folder(self):
        if self.output_folder and os.path.exists(self.output_folder):
            os.startfile(self.output_folder)
        else:
            messagebox.showerror("Lỗi", "Thư mục kết quả không tồn tại!")

    def on_api_key_focus_out(self):
        api_key = self.api_key_var.get().strip()
        if api_key:
            thread = threading.Thread(target=self.fetch_account_info, args=(api_key,))
            thread.daemon = True
            thread.start()
        else:
            self.root.after(0, lambda: self.account_info_frame.grid_remove())
    
    def fetch_account_info(self, api_key):
        try:
            api_domain = API_DOMAIN_TEST if api_key == TEST_API_KEY else API_DOMAIN_PROD
            account_status_endpoint = f"{api_domain}/api/account-status"
            
            response = requests.get(f"{account_status_endpoint}?key={api_key}", timeout=10)
            data = response.json()
            if response.status_code == 200:
                if data.get('success'):
                    account_data = data.get('data', {})
                    self.account_info['status'] = 'hoạt động'
                    self.account_info['balance'] = account_data.get('balance', 0)
                    self.account_info['createdAt'] = account_data.get('createdAt', '-')
                    self.account_info['updatedAt'] = account_data.get('updatedAt', '-')
                    
                    self.root.after(0, self.update_account_info_display)
                else:
                    self.account_info['status'] = data.get('error', 'không xác định')
                    self.account_info['balance'] = -1
                    self.account_info['createdAt'] = "-"
                    self.account_info['updatedAt'] = "-"
                self.root.after(0, self.update_account_info_display)
            else:
                self.account_info['status'] = data.get('error', 'không xác định')
                self.account_info['balance'] = -1
                self.account_info['createdAt'] = "-"
                self.account_info['updatedAt'] = "-"
                self.root.after(0, self.update_account_info_display)
        except Exception as e:
            self.account_info['status'] = f"không xác định: ({str(e)})"
            self.account_info['balance'] = -1
            self.account_info['createdAt'] = "-"
            self.account_info['updatedAt'] = "-"
            self.root.after(0, self.update_account_info_display)
    
    def update_account_info_display(self):
        from datetime import datetime
        
        self.account_name_label.configure(
            text=f"Tình trạng: {self.account_info['status']}"
        )
        
        balance_formatted = f"{self.account_info['balance']:,}đ".replace(',', '.')
        self.account_balance_label.configure(
            text=f"Số dư: {balance_formatted}"
        )
        
        try:
            created_dt = datetime.fromisoformat(self.account_info['createdAt'].replace('Z', '+00:00'))
            created_str = created_dt.strftime('%d/%m/%Y')
        except Exception:
            created_str = '--'
        
        self.account_created_label.configure(
            text=f"Ngày tạo: {created_str}"
        )
        
        try:
            updated_dt = datetime.fromisoformat(self.account_info['updatedAt'].replace('Z', '+00:00'))
            updated_str = updated_dt.strftime('%d/%m/%Y %H:%M')
        except Exception:
            updated_str = '--'
        
        self.account_updated_label.configure(
            text=f"Cập nhật: {updated_str}"
        )
        
        self.account_info_frame.grid()

    def toggle_api_key_visibility(self):
        self.api_key_visible = not self.api_key_visible
        
        if self.api_key_visible:
            self.api_entry.configure(show="")
            self.toggle_api_btn.configure(text="Ẩn")
        else:
            self.api_entry.configure(show="•")
            self.toggle_api_btn.configure(text="Hiện")


def main():
    app = AddressConverterApp()
    app.root.mainloop()


if __name__ == "__main__":
    main()
