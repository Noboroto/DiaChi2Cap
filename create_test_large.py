"""
Script tạo file test với > 1000 địa chỉ để test auto-batch
"""
import openpyxl
from openpyxl.styles import Font, Alignment

test_addresses = [
    "470 Đ. Trần Đại Nghĩa, Hoà Hải, Ngũ Hành Sơn, Đà Nẵng",
    "54 Nguyễn Lương Bằng, Hoà Khánh Bắc, Liên Chiểu, Đà Nẵng",
    "456 đường Lương Định Của, Phường An Khánh, Quận 2, TP. Hồ Chí Minh",
    "123 đường Man Thiện, Phường Tăng Nhơn Phú A, Quận 9, TP. Hồ Chí Minh"
]

def create_large_txt():
    """Tạo file txt với 1500 địa chỉ"""
    with open('test_input_large.txt', 'w', encoding='utf-8') as f:
        for i in range(1500):
            addr = test_addresses[i % len(test_addresses)]
            f.write(f"{addr}\n")
    print("[OK] Đã tạo test_input_large.txt với 1500 địa chỉ")

def create_large_excel():
    """Tạo file Excel với 2500 địa chỉ"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Địa chỉ"
    
    ws['A1'] = "Địa chỉ"
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.column_dimensions['A'].width = 60
    
    for i in range(2500):
        addr = test_addresses[i % len(test_addresses)]
        ws[f'A{i+2}'] = addr
    
    wb.save('test_input_large.xlsx')
    wb.close()
    print("[OK] Đã tạo test_input_large.xlsx với 2500 địa chỉ")

if __name__ == "__main__":
    print("Đang tạo file test lớn...")
    print()
    
    create_large_txt()
    create_large_excel()
    
    print()
    print("[HOÀN THÀNH] Sử dụng các file này để test auto-batch:")
    print("  - test_input_large.txt: 1500 địa chỉ (2 batch)")
    print("  - test_input_large.xlsx: 2500 địa chỉ (3 batch)")
