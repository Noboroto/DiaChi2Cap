# -*- coding: utf-8 -*-
"""
File I/O handlers for reading and writing address data
Supports both .txt and .xlsx formats
"""
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from modules.code_lookup import extract_codes_from_address


def read_txt_file(file_path):
    """
    Read txt file and detect format: single column or multi-column (5+ columns)
    
    Args:
        file_path: Path to txt file
        
    Returns:
        tuple: (data, is_multi_column, headers)
            - data: List of addresses or list of dicts for multi-column
            - is_multi_column: Boolean indicating format type
            - headers: List of headers for multi-column, None for single column
    """
    data = []
    is_multi_column = False
    headers = None
    
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = [line.strip() for line in f if line.strip()]
    
    if not lines:
        return [], False, None
    
    first_line_parts = [p.strip() for p in lines[0].split(',')]
    
    if len(first_line_parts) >= 5:
        is_multi_column = True
        headers = first_line_parts
        
        for line in lines[1:]:
            parts = [p.strip() for p in line.split(',')]
            if len(parts) >= 5:
                row_data = {
                    'code': parts[0],
                    'so_nha_duong': parts[1],
                    'phuong_xa': parts[2],
                    'quan_huyen': parts[3],
                    'tinh_thanh': parts[4],
                    'extra_columns': parts[5:] if len(parts) > 5 else []
                }
                data.append(row_data)
    else:
        for line in lines:
            if line:
                data.append(line)
    
    return data, is_multi_column, headers


def read_excel_file(file_path):
    """
    Read Excel file and detect format: single column or multi-column (5+ columns)
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        tuple: (data, is_multi_column, headers)
            - data: List of addresses or list of dicts for multi-column
            - is_multi_column: Boolean indicating format type
            - headers: List of headers for multi-column, None for single column
    """
    data = []
    is_multi_column = False
    headers = None
    
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    
    if ws is None:
        return [], False, None
    
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    
    if not rows:
        return [], False, None
    
    first_row = rows[0]
    num_cols = sum(1 for cell in first_row if cell is not None and str(cell).strip())
    
    if num_cols >= 5:
        is_multi_column = True
        headers = [str(cell).strip() if cell is not None else '' for cell in first_row]
        
        for row in rows[1:]:
            if row and any(cell is not None for cell in row):
                cells = [str(cell).strip() if cell is not None else '' for cell in row]
                if len(cells) >= 5 and any(cells[:5]):
                    row_data = {
                        'code': cells[0] if len(cells) > 0 else '',
                        'so_nha_duong': cells[1] if len(cells) > 1 else '',
                        'phuong_xa': cells[2] if len(cells) > 2 else '',
                        'quan_huyen': cells[3] if len(cells) > 3 else '',
                        'tinh_thanh': cells[4] if len(cells) > 4 else '',
                        'extra_columns': cells[5:] if len(cells) > 5 else []
                    }
                    data.append(row_data)
    else:
        for row in rows:
            if row and row[0]:
                addr = str(row[0]).strip()
                if addr:
                    data.append(addr)
    
    return data, is_multi_column, headers


def write_txt_output(output_path, results, is_multi_column=False, headers=None, original_data=None):
    """
    Write output to txt file
    
    Args:
        output_path: Path to output file
        results: List of conversion results
        is_multi_column: Boolean indicating if multi-column format
        headers: List of header names (for multi-column)
        original_data: List of original row dicts (for multi-column)
    """
    with open(output_path, 'w', encoding='utf-8') as f:
        if is_multi_column and headers and original_data:
            new_headers = headers + ['Số lần thử', 'Mã phường/xã mới', 'Phường/Xã mới', 'Mã tỉnh/thành mới', 'Tỉnh/Thành mới']
            f.write(','.join(new_headers) + '\n')
            
            for idx, result in enumerate(results):
                if idx < len(original_data):
                    row_data = original_data[idx]
                    row_parts = [
                        row_data.get('code', ''),
                        row_data.get('so_nha_duong', ''),
                        row_data.get('phuong_xa', ''),
                        row_data.get('quan_huyen', ''),
                        row_data.get('tinh_thanh', '')
                    ]
                    row_parts.extend(row_data.get('extra_columns', []))
                    
                    retry_count = result.get('retryCount', 1)
                    row_parts.append(str(retry_count))
                    
                    if result.get("success", False):
                        converted = result.get("converted", "")
                        
                        province_code, ward_code = extract_codes_from_address(converted)
                        
                        parts = [p.strip() for p in converted.split(',')]
                        ward = parts[-2] if len(parts) > 0 else ""
                        province = parts[-1] if len(parts) > 0 else ""
                        
                        row_parts.append(ward_code or '')
                        row_parts.append(ward)
                        row_parts.append(province_code or '')
                        row_parts.append(province)
                    else:
                        error_msg = result.get("error", "Lỗi không xác định")
                        error_text = f"LỖI: {error_msg} - {result.get('original', '')}"
                        row_parts.append('')
                        row_parts.append(error_text)
                        row_parts.append('')
                        row_parts.append('')
                    
                    f.write(','.join(row_parts) + '\n')
        else:
            for result in results:
                original = result.get('original', '')
                if result.get("success", False):
                    f.write(result.get("converted", "") + "\n")
                else:
                    f.write(f"{original}\n")


def write_excel_output(output_path, results, is_multi_column=False, headers=None, original_data=None):
    """
    Write output to Excel file
    
    Args:
        output_path: Path to output file
        results: List of conversion results
        is_multi_column: Boolean indicating if multi-column format
        headers: List of header names (for multi-column)
        original_data: List of original row dicts (for multi-column)
    """
    wb = Workbook()
    ws = wb.active
    
    if ws is None:
        raise ValueError("Không thể tạo worksheet")
    
    ws.title = "Kết quả"
    
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    if is_multi_column and headers and original_data:
        new_headers = headers + ['Số lần thử', 'Mã phường/xã mới', 'Phường/Xã mới', 'Mã tỉnh/thành mới', 'Tỉnh/Thành mới']
        
        for col_idx, header in enumerate(new_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        for idx, result in enumerate(results, start=2):
            orig_idx = idx - 2
            if orig_idx < len(original_data):
                row_data = original_data[orig_idx]
                
                ws.cell(row=idx, column=1, value=row_data.get('code', ''))
                ws.cell(row=idx, column=2, value=row_data.get('so_nha_duong', ''))
                ws.cell(row=idx, column=3, value=row_data.get('phuong_xa', ''))
                ws.cell(row=idx, column=4, value=row_data.get('quan_huyen', ''))
                ws.cell(row=idx, column=5, value=row_data.get('tinh_thanh', ''))
                
                extra_cols = row_data.get('extra_columns', [])
                for extra_idx, extra_val in enumerate(extra_cols):
                    ws.cell(row=idx, column=6+extra_idx, value=extra_val)
                
                retry_col = 6 + len(extra_cols)
                ward_code_col = retry_col + 1
                ward_col = ward_code_col + 1
                province_code_col = ward_col + 1
                province_col = province_code_col + 1
                
                retry_count = result.get('retryCount', 1)
                ws.cell(row=idx, column=retry_col, value=retry_count)
                
                if result.get("success", False):
                    converted = result.get("converted", "")
                    
                    province_code, ward_code = extract_codes_from_address(converted)
                    
                    parts = [p.strip() for p in converted.split(',')]
                    ward = parts[-2] if len(parts) > 0 else ""
                    province = parts[-1] if len(parts) > 0 else ""
                    
                    if "[LỖI]" in ward:
                        ward = ward.replace("[LỖI]", "LỖI:").strip()
                        cell = ws.cell(row=idx, column=ward_code_col, value=ward)
                        cell.font = Font(color="FF0000")
                        ws.cell(row=idx, column=ward_col, value='')
                        ws.cell(row=idx, column=province_code_col, value='')
                        ws.cell(row=idx, column=province_col, value='')
                        continue

                    ws.cell(row=idx, column=ward_code_col, value=ward_code or '')
                    ws.cell(row=idx, column=ward_col, value=ward)
                    ws.cell(row=idx, column=province_code_col, value=province_code or '')
                    ws.cell(row=idx, column=province_col, value=province)
                else:
                    error_msg = result.get("error", "Lỗi không xác định")
                    cell_code = ws.cell(row=idx, column=ward_code_col, value=f"LỖI: {error_msg}")
                    cell_code.font = Font(color="FF0000")
                    ws.cell(row=idx, column=ward_col, value='')
                    cell_value = ws.cell(row=idx, column=province_code_col, value=f"{result.get('original', '-')}")
                    cell_value.font = Font(color="FF0000")
                    ws.cell(row=idx, column=province_col, value='')
                    

        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(new_headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
    else:
        ws['A1'] = "Gốc"
        ws['B1'] = "Chuyển đổi"
        
        ws['A1'].font = header_font
        ws['B1'].font = header_font
        ws['A1'].fill = header_fill
        ws['B1'].fill = header_fill
        ws['A1'].alignment = header_alignment
        ws['B1'].alignment = header_alignment
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 60
        
        for idx, result in enumerate(results, start=2):
            ws[f'A{idx}'] = result.get("original", "")
            if result.get("success", False):
                ws[f'B{idx}'] = result.get("converted", "")
            else:
                error_msg = result.get("error", "Lỗi không xác định")
                ws[f'B{idx}'] = f"LỖI: {error_msg};{result.get('original', '')}"
                ws[f'B{idx}'].font = Font(color="FF0000")
    
    wb.save(output_path)
    wb.close()
