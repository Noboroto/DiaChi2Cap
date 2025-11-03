"""
Script tạo file Excel test
"""
import openpyxl
from openpyxl import Workbook

def create_test_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Địa chỉ"
    
    ws['A1'] = "Địa chỉ"
    
    test_addresses = [
        "470 Đ. Trần Đại Nghĩa, Hoà Hải, Ngũ Hành Sơn, Đà Nẵng",
        "54 Nguyễn Lương Bằng, Hoà Khánh Bắc, Liên Chiểu, Đà Nẵng",
        "123 đường Man Thiện, Phường Tăng Nhơn Phú A, Quận 9, TP. Hồ Chí Minh.",
        "456 đường Lương Định Của, Phường An Khánh, Quận 2, TP. Hồ Chí Minh.",
        "470 Trần Đại Nghĩa, Hoà Hải, Ngũ Hành Sơn, Đà Nẵng"
    ]
    
    for idx, addr in enumerate(test_addresses, start=2):
        ws[f'A{idx}'] = addr
    
    ws.column_dimensions['A'].width = 70
    
    wb.save("test_input.xlsx")
    print("Đã tạo file test_input.xlsx")

if __name__ == "__main__":
    create_test_excel()
